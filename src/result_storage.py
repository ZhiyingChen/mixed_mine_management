from typing import List
import openpyxl
import logging
import pandas as pd

from .input_data import InputData
from .utils import field, header, functions


class ResultStorage:
    def __init__(
            self,
            input_data: InputData,
            keys: List[str],
            result,
            multi_results
    ):
        self.input_data: InputData = input_data
        self.keys = keys
        self.result = result
        self.multi_results = multi_results

    def write_to_excel(self):
        if not self.result.success:
            logging.error(
                "Unsuccessful solution, message: {}".format(self.result.message)
            )
        else:
            logging.info("Successful solution, message: {}".format(self.result.message))

        sh = header.MaterialHeader
        # 加载当前工作簿
        file_name = "{}{}".format(self.input_data.exe_folder, field.ROCK_FILENAME)
        wb = openpyxl.load_workbook(file_name)

        # 选择指定的 sheet
        sheet = wb[field.MATERIAL_SHEET]  # 替换为你的 sheet 名
        sheet_header_dict = functions.get_header_dict(sheet=sheet)
        # 要输出的数据
        data_to_write = dict(zip(self.keys, self.result.x))
        data_key = [
            cell.value
            for cell in sheet["A"]
            if cell.value is not None and cell.value != sh.material_name
        ]
        data_val = [data_to_write[key] for key in data_key]
        # 替换为你要写入的数据
        # 指定要写入的列（例如，第2列）
        column_index = sheet_header_dict[sh.ratio]
        # A=1, B=2, C=3, ...
        # 遍历每一行
        for row_index, value in enumerate(data_val, start=2):  # 从第1行开始
            sheet.cell(row=row_index, column=column_index, value=value)
        # 保存工作簿
        wb.save(file_name)

    def generate_multi_results(self):
        data = {material_name: [] for material_name in self.keys}
        data.update({"原材料成本": []})
        for idx, (result_ratio, result_obj) in enumerate(self.multi_results):
            for k, material_name in enumerate(self.keys):
                data[material_name].append(result_ratio[k])
            data['原材料成本'].append(result_obj)

    def write_multi_results_to_excel(self):
        rh = header.MultiResultHeader
        # 构建列名（如：["材料名称", "结果0配比", "结果1配比"]）
        col = [rh.material_name] + [f'结果{i}配比' for i in range(len(self.multi_results))]

        # 准备数据
        data = {material: [] for material in self.keys}
        cost_list = []
        for result_ratio, result_obj in self.multi_results:
            for i, material in enumerate(self.keys):
                data[material].append(result_ratio[i])
            cost_list.append(result_obj)

        # 构建记录列表（DataFrame 数据）
        records = []
        for material in self.keys:
            record = {rh.material_name: material}
            for i, value in enumerate(data[material]):
                record[f'结果{i}配比'] = value
            records.append(record)
        records.append(dict())
        # 添加原材料成本行
        cost_record = {rh.material_name: '原材料成本'}
        for i, cost in enumerate(cost_list):
            cost_record[f'结果{i}配比'] = cost
        records.append(cost_record)

        # 转换为 DataFrame
        result_df = pd.DataFrame(records, columns=col)

        # 写入 Excel
        file_name = f"{self.input_data.exe_folder}{field.ROCK_FILENAME}"
        try:
            # 使用追加模式加载工作簿
            with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                # 如果工作表不存在则创建
                if field.MULTI_RESULTS_SHEET not in writer.book.sheetnames:
                    writer.book.create_sheet(field.MULTI_RESULTS_SHEET)
                startrow = 0
                # 写入数据（不带表头）
                result_df.to_excel(writer, sheet_name=field.MULTI_RESULTS_SHEET, index=False, startrow=startrow,
                                   header=True)
        except Exception as e:
            logging.error(f"写入 Excel 失败: {e}")
