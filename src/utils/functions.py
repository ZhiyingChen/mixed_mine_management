import os
import platform
from openpyxl import load_workbook


def get_header_dict(sheet):
    # 获取表头和对应列索引的字典
    header_dict = {}
    for column in range(1, sheet.max_column + 1):
        header = sheet.cell(row=1, column=column).value
        if header:
            header_dict[header] = column
    return header_dict
